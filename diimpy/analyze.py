# Copyright (c) 2025 Stig Rune Sellevag
#
# This file is distributed under the MIT License. See the accompanying file
# LICENSE.txt or http://www.opensource.org/licenses/mit-license.php for terms
# and conditions.

import numpy as np
import pandas as pd

from openpyxl import load_workbook


def analyze_dependency(model, sheet_name="Dependency"):
    """Analyze dependencies and influence gains."""
    #
    # Output is written to existing datafile (xslx).
    #
    wb = load_workbook(model.config["datafile"])
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    header = ["function", "delta", "delta_overall", "rho", "rho_overall"]
    ws.append(header)

    delta = model.dependency()
    delta_overall = model.overall_dependency()
    rho = model.influence()
    rho_overall = model.overall_influence()

    res = []
    for f, d, do, r, ro in zip(model.infra, delta, delta_overall, rho, rho_overall):
        row = [f, d, do, r, ro]
        ws.append(row)
        res.append(row)

    wb.save(filename=model.config["datafile"])
    return pd.DataFrame(res, columns=header)


def analyze_interdependency(model, sheet_name="Interdependency"):
    """Analyze first-, second- and third-order interdependencies."""
    #
    # Output is written to existing datafile (xslx).
    #
    wb = load_workbook(model.config["datafile"])
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    header = ["i", "j", "max(aij)", "i", "j", "max(aij^2)", "i", "j", "max(aij^3)"]
    ws.append(header)

    first = model.max_nth_order_interdependency(1)
    second = model.max_nth_order_interdependency(2)
    third = model.max_nth_order_interdependency(3)

    res = []
    for i in range(len(model)):
        row = [
            first[i][0],
            first[i][1],
            first[i][2],
            second[i][0],
            second[i][1],
            second[i][2],
            third[i][0],
            third[i][1],
            third[i][2],
        ]
        ws.append(row)
        res.append(row)

    wb.save(filename=model.config["datafile"])
    return pd.DataFrame(res, columns=header)


def analyze_inoperability(model, sheet_name="Static_inoperability"):
    """Analyze inoperabilities at equilibrium."""
    #
    # Output is written to existing datafile (xslx).
    #
    wb = load_workbook(model.config["datafile"])
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    header = ["infrastructure", "inoperability"]
    ws.append(header)

    q = model.inoperability()

    res = []
    for i in range(len(model)):
        row = [model.infra[i], q[i]]
        ws.append(row)
        res.append(row)

    wb.save(filename=model.config["datafile"])
    return pd.DataFrame(res, columns=header)


def analyze_dynamic(model, sheet_name="Dynamic_inoperability"):
    """Analyze dynamic inoperabilities."""
    #
    # Output is written to existing datafile (xslx).
    #
    wb = load_workbook(model.config["datafile"])
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    header = ["time"]
    for item in model.infra:
        header.append(item)
    ws.append(header)

    qt = model.dynamic_inoperability()
    qtot = model.impact(qt)

    res = []
    for i in range(np.size(qt, 0)):
        row = [qt[i, 0]]
        for j in range(1, np.size(qt, 1)):
            row.append(qt[i, j])
        ws.append(row)
        res.append(row)

    row = ["qtot"]
    for item in qtot:
        row.append(item)
    ws.append(row)

    wb.save(filename=model.config["datafile"])
    return pd.DataFrame(res, columns=header), qtot


def analyze_recovery(model, sheet_name="Recovery"):
    """Analyze dynamic recovery."""
    #
    # Output is written to existing datafile (xslx).
    #
    wb = load_workbook(model.config["datafile"])
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    header = ["time"]
    for item in model.infra:
        header.append(item)
    ws.append(header)

    qt = model.dynamic_recovery()
    qtot = model.impact(qt)

    res = []
    for i in range(np.size(qt, 0)):
        row = [qt[i, 0]]
        for j in range(1, np.size(qt, 1)):
            row.append(qt[i, j])
        ws.append(row)
        res.append(row)

    row = ["qtot"]
    for item in qtot:
        row.append(item)
    ws.append(row)

    wb.save(filename=model.config["datafile"])
    return pd.DataFrame(res, columns=header), qtot


def single_attack_sampling(model, sheet_name="Single_attack", ptime=None, cvalue=None):
    """Run DIIM single attack sampling."""
    #
    # Output is written to existing datafile (xslx).
    #
    wb = load_workbook(model.config["datafile"])
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    header = ["infra", "impact"]
    ws.append(header)

    res = []
    for item in model.infra:
        model.set_perturbation(pinfra=[item], ptime=ptime, cvalue=cvalue)
        qt = model.dynamic_inoperability()
        qtot = model.impact(qt)
        row = [item, np.sum(qtot)]
        ws.append(row)
        res.append(row)

    wb.save(filename=model.config["datafile"])
    return pd.DataFrame(res, columns=header)


def hybrid_attack_sampling(model, sheet_name="Hybrid_attack", ptime=None, cvalue=None):
    """Run DIIM hybrid attack sampling."""
    #
    # Output is written to existing datafile (xslx).
    #
    wb = load_workbook(model.config["datafile"])
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    header = ["infra_i", "infra_j", "impact"]
    ws.append(header)

    if ptime is None:
        ptime = [model.perturb.config["ptime"][0], model.perturb.config["ptime"][0]]
    if cvalue is None:
        cvalue = [model.perturb.config["cvalue"][0], model.perturb.config["cvalue"][0]]

    res = []
    for ii in model.infra:
        for jj in model.infra:
            if ii == jj:
                continue
            model.set_perturbation(pinfra=[ii, jj], ptime=ptime, cvalue=cvalue)
            qt = model.dynamic_inoperability()
            qtot = model.impact(qt)
            row = [ii, jj, np.sum(qtot)]
            ws.append(row)
            res.append(row)

    wb.save(filename=model.config["datafile"])
    return pd.DataFrame(res, columns=header)
